"""Data import service for CSV and Excel files."""
from __future__ import annotations

import csv
import io
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from ..data import crm_repository, order_repository, product_repository
from ..models.order_models import CRMContact, Order, Product


@dataclass
class ImportColumn:
    """Represents a column in the import file."""
    index: int
    name: str
    sample_values: List[str] = field(default_factory=list)


@dataclass
class ColumnMapping:
    """Maps a source column to a target field."""
    source_column: int  # Index in the source file
    target_field: str   # Field name in the target model
    transform: Optional[str] = None  # Optional transformation: "date", "number", "boolean"


@dataclass
class ImportResult:
    """Result of an import operation."""
    success: bool
    imported_count: int
    skipped_count: int
    error_count: int
    errors: List[str] = field(default_factory=list)
    warnings: List[str] = field(default_factory=list)


# Field definitions for different import types
PRODUCT_FIELDS: Dict[str, Dict[str, Any]] = {
    "sku": {"label": "SKU", "required": True, "type": "text"},
    "name": {"label": "Product Name", "required": True, "type": "text"},
    "description": {"label": "Description", "required": False, "type": "text"},
    "inventory_count": {"label": "Inventory Count", "required": False, "type": "number", "default": 0},
    "base_unit_cost": {"label": "Base Unit Cost", "required": False, "type": "number", "default": 0.0},
    "default_unit_price": {"label": "Default Price", "required": False, "type": "number", "default": 0.0},
    "status": {"label": "Status", "required": False, "type": "text", "default": "Active"},
}

ORDER_FIELDS: Dict[str, Dict[str, Any]] = {
    "order_number": {"label": "Order Number", "required": True, "type": "text"},
    "customer_name": {"label": "Customer Name", "required": True, "type": "text"},
    "customer_address": {"label": "Address", "required": False, "type": "text", "default": ""},
    "order_date": {"label": "Order Date", "required": True, "type": "date"},
    "status": {"label": "Status", "required": False, "type": "text", "default": "Pending"},
    "is_paid": {"label": "Paid", "required": False, "type": "boolean", "default": False},
    "carrier": {"label": "Carrier", "required": False, "type": "text", "default": ""},
    "tracking_number": {"label": "Tracking Number", "required": False, "type": "text", "default": ""},
    "notes": {"label": "Notes", "required": False, "type": "text", "default": ""},
}

CUSTOMER_FIELDS: Dict[str, Dict[str, Any]] = {
    "name": {"label": "Customer Name", "required": True, "type": "text"},
    "company": {"label": "Company", "required": False, "type": "text", "default": ""},
    "address": {"label": "Address", "required": False, "type": "text", "default": ""},
    "email": {"label": "Email", "required": False, "type": "text", "default": ""},
    "phone": {"label": "Phone", "required": False, "type": "text", "default": ""},
    "tags": {"label": "Tags", "required": False, "type": "text", "default": ""},
    "notes": {"label": "Notes", "required": False, "type": "text", "default": ""},
}


def _read_all_rows(file_path: str) -> List[List[str]]:
    """Read every data row using the same format rules as the preview."""
    path = Path(file_path)
    suffix = path.suffix.lower()
    if suffix == ".csv":
        _, _, encoding = read_csv_preview(file_path, max_rows=1)
        with path.open("r", encoding=encoding, newline="") as handle:
            rows = list(csv.reader(handle))
        return rows[1:]
    if suffix == ".xlsx":
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise ImportError("Excel import requires the 'openpyxl' package") from exc
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            sheet = workbook.active
            if sheet is None:
                raise ValueError("No active sheet found in workbook")
            rows = [[str(cell) if cell is not None else "" for cell in row] for row in sheet.iter_rows(values_only=True)]
        finally:
            workbook.close()
        return rows[1:]
    if suffix == ".xls":
        raise ValueError("Legacy .xls files are not supported. Save the workbook as .xlsx and try again.")
    raise ValueError(f"Unsupported file format: {suffix}")


def read_csv_preview(file_path: str, max_rows: int = 10) -> Tuple[List[ImportColumn], List[List[str]], str]:
    """
    Read a preview of a CSV file.
    
    Args:
        file_path: Path to the CSV file
        max_rows: Maximum number of data rows to read for preview
        
    Returns:
        Tuple of (columns, data_rows, encoding_used)
    """
    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"File not found: {file_path}")
    
    # Try different encodings
    encodings = ["utf-8-sig", "utf-8", "cp1252", "latin-1"]
    content: str = ""
    encoding_used: str = "utf-8"
    
    for encoding in encodings:
        try:
            content = path.read_text(encoding=encoding)
            encoding_used = encoding
            break
        except UnicodeDecodeError:
            continue
    
    if not content:
        raise ValueError("Could not decode file with any supported encoding")
    
    # Parse CSV
    if not content.strip():
        raise ValueError("File is empty")

    reader = csv.reader(io.StringIO(content, newline=""))
    rows = list(reader)
    
    if not rows:
        raise ValueError("No data found in file")
    
    # First row is headers
    headers = rows[0]
    data_rows = rows[1:max_rows + 1]
    
    # Build column info with sample values
    columns: List[ImportColumn] = []
    for i, header in enumerate(headers):
        samples = [row[i] if i < len(row) else "" for row in data_rows[:5]]
        columns.append(ImportColumn(
            index=i,
            name=header.strip() or f"Column {i + 1}",
            sample_values=samples,
        ))
    
    return columns, data_rows, encoding_used


def read_excel_preview(file_path: str, max_rows: int = 10) -> Tuple[List[ImportColumn], List[List[str]], str]:
    """
    Read a preview of an Excel file.
    
    Requires openpyxl to be installed.
    
    Args:
        file_path: Path to the Excel file
        max_rows: Maximum number of data rows to read for preview
        
    Returns:
        Tuple of (columns, data_rows, sheet_name)
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ImportError("Excel import requires the 'openpyxl' package. Install it with: pip install openpyxl")
    
    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"File not found: {file_path}")
    
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    if sheet is None:
        raise ValueError("No active sheet found in workbook")
    
    sheet_name = sheet.title
    
    rows: List[List[str]] = []
    for row in sheet.iter_rows(max_row=max_rows + 1, values_only=True):
        row_values = [str(cell) if cell is not None else "" for cell in row]
        rows.append(row_values)
    
    workbook.close()
    
    if not rows:
        raise ValueError("No data found in file")
    
    # First row is headers
    headers = rows[0]
    data_rows = rows[1:]
    
    # Build column info with sample values
    columns: List[ImportColumn] = []
    for i, header in enumerate(headers):
        samples = [row[i] if i < len(row) else "" for row in data_rows[:5]]
        columns.append(ImportColumn(
            index=i,
            name=str(header).strip() or f"Column {i + 1}",
            sample_values=samples,
        ))
    
    return columns, data_rows, sheet_name


def auto_map_columns(
    source_columns: List[ImportColumn],
    target_fields: Dict[str, Dict[str, Any]],
) -> List[ColumnMapping]:
    """
    Automatically map source columns to target fields based on name similarity.
    
    Args:
        source_columns: Columns from the import file
        target_fields: Target field definitions
        
    Returns:
        List of column mappings
    """
    mappings: List[ColumnMapping] = []
    
    # Build lookup of normalized names
    field_lookup: Dict[str, str] = {}
    for field_name, field_def in target_fields.items():
        # Add field name variations
        normalized = field_name.lower().replace("_", " ").strip()
        field_lookup[normalized] = field_name
        
        # Add label variations
        label = field_def.get("label", "").lower().strip()
        if label:
            field_lookup[label] = field_name
    
    used_fields: set = set()
    
    for col in source_columns:
        normalized_name = col.name.lower().replace("_", " ").strip()
        
        # Direct match
        if normalized_name in field_lookup:
            target = field_lookup[normalized_name]
            if target not in used_fields:
                field_type = target_fields[target].get("type", "text")
                transform = None
                if field_type == "date":
                    transform = "date"
                elif field_type == "number":
                    transform = "number"
                elif field_type == "boolean":
                    transform = "boolean"
                
                mappings.append(ColumnMapping(
                    source_column=col.index,
                    target_field=target,
                    transform=transform,
                ))
                used_fields.add(target)
                continue
        
        # Partial match
        for lookup_name, field_name in field_lookup.items():
            if field_name in used_fields:
                continue
            if lookup_name in normalized_name or normalized_name in lookup_name:
                field_type = target_fields[field_name].get("type", "text")
                transform = None
                if field_type == "date":
                    transform = "date"
                elif field_type == "number":
                    transform = "number"
                elif field_type == "boolean":
                    transform = "boolean"
                
                mappings.append(ColumnMapping(
                    source_column=col.index,
                    target_field=field_name,
                    transform=transform,
                ))
                used_fields.add(field_name)
                break
    
    return mappings


def _parse_date(value: str) -> Optional[date]:
    """Parse a date string into a date object."""
    if not value or value.lower() in ("none", "null", ""):
        return None
    
    # Try common date formats
    formats = [
        "%Y-%m-%d",
        "%m/%d/%Y",
        "%m-%d-%Y",
        "%d/%m/%Y",
        "%d-%m-%Y",
        "%Y/%m/%d",
        "%m/%d/%y",
        "%d/%m/%y",
    ]
    
    for fmt in formats:
        try:
            return datetime.strptime(value.strip(), fmt).date()
        except ValueError:
            continue
    
    return None


def _parse_number(value: str) -> float:
    """Parse a number string into a float."""
    if not value or value.lower() in ("none", "null", ""):
        return 0.0
    
    # Remove currency symbols and commas
    cleaned = value.replace("$", "").replace(",", "").replace(" ", "").strip()
    
    # Handle parentheses for negative numbers
    if cleaned.startswith("(") and cleaned.endswith(")"):
        cleaned = "-" + cleaned[1:-1]
    
    try:
        return float(cleaned)
    except ValueError:
        return 0.0


def _parse_boolean(value: str) -> bool:
    """Parse a boolean string."""
    if not value:
        return False
    
    normalized = value.lower().strip()
    return normalized in ("true", "yes", "1", "y", "x", "✓", "checked")


def import_products(
    file_path: str,
    mappings: List[ColumnMapping],
    skip_duplicates: bool = True,
) -> ImportResult:
    """
    Import products from a file.
    
    Args:
        file_path: Path to the import file
        mappings: Column mappings
        skip_duplicates: If True, skip products with existing SKUs; if False, update them
        
    Returns:
        ImportResult with counts and errors
    """
    try:
        all_rows = _read_all_rows(file_path)
    except (ImportError, OSError, ValueError) as exc:
        return ImportResult(
            success=False,
            imported_count=0,
            skipped_count=0,
            error_count=1,
            errors=[str(exc)],
        )
    
    # Build field mapping lookup
    mapping_lookup: Dict[str, ColumnMapping] = {m.target_field: m for m in mappings}
    
    imported = 0
    skipped = 0
    errors: List[str] = []
    warnings: List[str] = []
    
    for row_idx, row in enumerate(all_rows, start=2):
        try:
            # Extract SKU (required)
            sku_mapping = mapping_lookup.get("sku")
            if not sku_mapping:
                errors.append(f"Row {row_idx}: SKU mapping required")
                continue
            
            sku = row[sku_mapping.source_column].strip().upper() if sku_mapping.source_column < len(row) else ""
            if not sku:
                errors.append(f"Row {row_idx}: Empty SKU")
                continue
            
            # Check for existing product
            existing = product_repository.get_product_by_sku(sku)
            if existing and skip_duplicates:
                skipped += 1
                warnings.append(f"Row {row_idx}: SKU '{sku}' already exists, skipped")
                continue
            
            # Build product data
            def get_value(field: str) -> str:
                mapping = mapping_lookup.get(field)
                if not mapping or mapping.source_column >= len(row):
                    return ""
                val = row[mapping.source_column]
                return str(val).strip() if val is not None else ""
            
            def get_transformed(field: str, default: Any = None) -> Any:
                mapping = mapping_lookup.get(field)
                if not mapping or mapping.source_column >= len(row):
                    return default
                
                val = row[mapping.source_column]
                value = str(val).strip() if val is not None else ""
                if not value:
                    return default
                
                if mapping.transform == "number":
                    return _parse_number(value)
                elif mapping.transform == "boolean":
                    return _parse_boolean(value)
                
                return value

            name = get_value("name")
            if not name:
                errors.append(f"Row {row_idx}: Empty product name")
                continue

            product = Product(
                id=existing.id if existing else None,
                sku=sku,
                name=name,
                description=get_value("description") if "description" in mapping_lookup else (existing.description if existing else ""),
                photo_path=existing.photo_path if existing else "",
                inventory_count=int(get_transformed("inventory_count", existing.inventory_count if existing else 0)),
                is_complete=existing.is_complete if existing else False,
                status=(get_value("status") or (existing.status if existing else "Active")),
                base_unit_cost=float(get_transformed("base_unit_cost", existing.base_unit_cost if existing else 0.0)),
                default_unit_price=float(get_transformed("default_unit_price", existing.default_unit_price if existing else 0.0)),
                pricing_components=list(existing.pricing_components) if existing else [],
            )
            
            if existing:
                product_repository.update_product(product)
            else:
                # Use create_product and then update with full data
                created = product_repository.create_product(sku, product.name)
                product = Product(
                    id=created.id,
                    sku=sku,
                    name=product.name,
                    description=product.description,
                    photo_path="",
                    inventory_count=product.inventory_count,
                    is_complete=False,
                    status=product.status,
                    base_unit_cost=product.base_unit_cost,
                    default_unit_price=product.default_unit_price,
                    pricing_components=[],
                )
                product_repository.update_product(product)
            
            imported += 1
            
        except Exception as e:
            errors.append(f"Row {row_idx}: {str(e)}")
    
    return ImportResult(
        success=len(errors) == 0,
        imported_count=imported,
        skipped_count=skipped,
        error_count=len(errors),
        errors=errors,
        warnings=warnings,
    )


def import_orders(
    file_path: str,
    mappings: List[ColumnMapping],
    skip_duplicates: bool = True,
) -> ImportResult:
    """
    Import orders from a file.
    
    Args:
        file_path: Path to the import file
        mappings: Column mappings
        skip_duplicates: If True, skip orders with existing order numbers
        
    Returns:
        ImportResult with counts and errors
    """
    try:
        all_rows = _read_all_rows(file_path)
    except (ImportError, OSError, ValueError) as exc:
        return ImportResult(
            success=False,
            imported_count=0,
            skipped_count=0,
            error_count=1,
            errors=[str(exc)],
        )
    
    mapping_lookup: Dict[str, ColumnMapping] = {m.target_field: m for m in mappings}
    
    imported = 0
    skipped = 0
    errors: List[str] = []
    warnings: List[str] = []
    
    for row_idx, row in enumerate(all_rows, start=2):
        try:
            # Extract order number (required)
            order_num_mapping = mapping_lookup.get("order_number")
            if not order_num_mapping:
                errors.append(f"Row {row_idx}: Order number mapping required")
                continue
            
            raw_val = row[order_num_mapping.source_column] if order_num_mapping.source_column < len(row) else ""
            order_number = str(raw_val).strip() if raw_val is not None else ""
            if not order_number:
                errors.append(f"Row {row_idx}: Empty order number")
                continue
            
            existing = order_repository.fetch_order_by_number(order_number)
            if existing and skip_duplicates:
                skipped += 1
                warnings.append(f"Row {row_idx}: Order '{order_number}' already exists, skipped")
                continue
            
            def get_value(field: str) -> str:
                mapping = mapping_lookup.get(field)
                if not mapping or mapping.source_column >= len(row):
                    return ""
                val = row[mapping.source_column]
                return str(val).strip() if val is not None else ""
            
            def get_transformed(field: str, default: Any = None) -> Any:
                mapping = mapping_lookup.get(field)
                if not mapping or mapping.source_column >= len(row):
                    return default
                
                val = row[mapping.source_column]
                value = str(val).strip() if val is not None else ""
                if not value:
                    return default
                
                if mapping.transform == "date":
                    return _parse_date(value) or default
                elif mapping.transform == "number":
                    return _parse_number(value)
                elif mapping.transform == "boolean":
                    return _parse_boolean(value)
                
                return value
            
            customer_name = get_value("customer_name")
            if not customer_name:
                errors.append(f"Row {row_idx}: Empty customer name")
                continue
            order_date = _parse_date(get_value("order_date"))
            if order_date is None:
                errors.append(f"Row {row_idx}: Invalid order date")
                continue
            
            order = Order(
                id=existing.id if existing else None,
                order_number=order_number,
                customer_name=customer_name,
                customer_address=get_value("customer_address") if "customer_address" in mapping_lookup else (existing.customer_address if existing else ""),
                order_date=order_date,
                status=get_value("status") or (existing.status if existing else "Pending"),
                is_paid=bool(get_transformed("is_paid", existing.is_paid if existing else False)),
                carrier=get_value("carrier") if "carrier" in mapping_lookup else (existing.carrier if existing else ""),
                tracking_number=get_value("tracking_number") if "tracking_number" in mapping_lookup else (existing.tracking_number if existing else ""),
                notes=get_value("notes") if "notes" in mapping_lookup else (existing.notes if existing else ""),
                ship_date=existing.ship_date if existing else None,
                target_completion_date=existing.target_completion_date if existing else None,
                items=list(existing.items) if existing else [],
                tax_rate=existing.tax_rate if existing else 0.0,
                tax_amount=existing.tax_amount if existing else 0.0,
                tax_included_in_total=existing.tax_included_in_total if existing else False,
            )
            
            if existing and existing.id is not None:
                order_repository.update_order(existing.id, order)
            else:
                order_repository.insert_order(order)
            
            imported += 1
            
        except Exception as e:
            errors.append(f"Row {row_idx}: {str(e)}")
    
    return ImportResult(
        success=len(errors) == 0,
        imported_count=imported,
        skipped_count=skipped,
        error_count=len(errors),
        errors=errors,
        warnings=warnings,
    )


def import_customers(
    file_path: str,
    mappings: List[ColumnMapping],
    skip_duplicates: bool = True,
) -> ImportResult:
    """Import CRM contacts, matching duplicates by email first and then name."""
    try:
        all_rows = _read_all_rows(file_path)
    except (ImportError, OSError, ValueError) as exc:
        return ImportResult(False, 0, 0, 1, errors=[str(exc)])

    mapping_lookup = {mapping.target_field: mapping for mapping in mappings}
    existing_contacts = crm_repository.list_contacts()
    by_email = {contact.email.strip().casefold(): contact for contact in existing_contacts if contact.email.strip()}
    by_name = {contact.customer_name.strip().casefold(): contact for contact in existing_contacts if contact.customer_name.strip()}
    imported = 0
    skipped = 0
    errors: List[str] = []
    warnings: List[str] = []

    for row_idx, row in enumerate(all_rows, start=2):
        try:
            def get_value(field: str) -> str:
                mapping = mapping_lookup.get(field)
                if not mapping or mapping.source_column >= len(row):
                    return ""
                value = row[mapping.source_column]
                return str(value).strip() if value is not None else ""

            name = get_value("name")
            if not name:
                errors.append(f"Row {row_idx}: Empty customer name")
                continue
            email = get_value("email")
            existing = by_email.get(email.casefold()) if email else None
            existing = existing or by_name.get(name.casefold())
            if existing and skip_duplicates:
                skipped += 1
                warnings.append(f"Row {row_idx}: Customer '{name}' already exists, skipped")
                continue
            tags = [tag.strip() for tag in get_value("tags").replace(";", ",").split(",") if tag.strip()]
            contact = CRMContact(
                id=existing.id if existing else None,
                customer_name=name,
                company=get_value("company") if "company" in mapping_lookup else (existing.company if existing else ""),
                email=email if "email" in mapping_lookup else (existing.email if existing else ""),
                phone=get_value("phone") if "phone" in mapping_lookup else (existing.phone if existing else ""),
                address=get_value("address") if "address" in mapping_lookup else (existing.address if existing else ""),
                tags=tags if "tags" in mapping_lookup else (list(existing.tags) if existing else []),
                created_at=existing.created_at if existing else None,
                last_contacted=existing.last_contacted if existing else None,
                next_follow_up=existing.next_follow_up if existing else None,
                preferred_channel=existing.preferred_channel if existing else "",
                notes=get_value("notes") if "notes" in mapping_lookup else (existing.notes if existing else ""),
            )
            contact_id = crm_repository.save_contact(contact)
            contact.id = contact_id
            by_name[name.casefold()] = contact
            if contact.email:
                by_email[contact.email.casefold()] = contact
            imported += 1
        except Exception as exc:
            errors.append(f"Row {row_idx}: {exc}")

    return ImportResult(not errors, imported, skipped, len(errors), errors=errors, warnings=warnings)


def get_field_definitions(import_type: str) -> Dict[str, Dict[str, Any]]:
    """Get field definitions for an import type."""
    if import_type == "products":
        return PRODUCT_FIELDS
    elif import_type == "orders":
        return ORDER_FIELDS
    elif import_type == "customers":
        return CUSTOMER_FIELDS
    else:
        return {}
